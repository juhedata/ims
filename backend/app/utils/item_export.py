"""库存/快照单品明细导出 Excel 的共享工具。"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.enums import OutboundType
from app.utils.ledger_display import (
    LEDGER_INBOUND_DISPLAY_HEADERS,
    LEDGER_OUTBOUND_DISPLAY_HEADERS,
    inbound_display_values,
    outbound_display_values,
)

# 出库类库存状态 → 库存属性展示文案
_OUT_STOCK_CONDITION_LABELS = {
    "SOLD": "售出-线上",
    "SOLD_OFFLINE": "售出-线下",
    "PRESOLD": "准售出",
    "BORROWED": "已借用",
    "GIFTED": "赠送",
    "SCRAPPED": "损毁",
    "RND": "研发",
    "SAMPLE": "样机",
    "TRIAL": "试用",
    "REPAIR": "维修",
    "DEPT_PROCUREMENT": "部门采购",
}

_IN_STOCK_CONDITION_LABELS = {
    "NEW": "正常",
    "RETURNED_FROM_SALE": "线上售出退回",
    "RETURNED_FROM_SOLD_OFFLINE": "线下售出退回",
    "RETURNED_FROM_PRESOLD": "准售出退回",
    "RETURNED_FROM_BORROW": "借用归还",
    "RETURNED_FROM_GIFT": "赠送退回",
    "RETURNED_FROM_SCRAPPED": "损毁退回",
    "RETURNED_FROM_RND": "研发退回",
    "RETURNED_FROM_SAMPLE": "样机退回",
    "RETURNED_FROM_TRIAL": "试用退回",
    "RETURNED_FROM_REPAIR": "维修退回",
    "RETURNED_FROM_DEPT_PROCUREMENT": "部门采购退回",
}

_OPERATION_STATUS_LABELS = {
    "INITIATED": "已发起",
    "PICKING": "出库处理中",
    "CHECKING": "入库处理中",
    "COMPLETED": "已完成",
    "CANCELLED": "已取消",
    "FAILED": "失败",
}

_OUTBOUND_TYPE_LABELS = {
    OutboundType.SOLD.value: "售出-线上",
    OutboundType.SOLD_OFFLINE.value: "售出-线下",
    OutboundType.PRESOLD.value: "准售出",
    OutboundType.GIFTED.value: "赠送",
    OutboundType.SCRAPPED.value: "损毁",
    OutboundType.RND.value: "研发",
    OutboundType.SAMPLE.value: "样机",
    OutboundType.TRIAL.value: "试用",
    OutboundType.REPAIR.value: "维修",
    OutboundType.DEPT_PROCUREMENT.value: "部门采购",
    OutboundType.BORROWED.value: "借用",
}


def stock_status_label(stock_status: str | None) -> str:
    return "在库" if stock_status == "IN_STOCK" else "出库"


def stock_condition_label(stock_status: str | None, stock_condition: str | None) -> str:
    if stock_status == "IN_STOCK":
        return _IN_STOCK_CONDITION_LABELS.get(stock_condition, stock_condition or "在库")
    return _OUT_STOCK_CONDITION_LABELS.get(stock_status, stock_status or "")


def operation_status_label(operation_status: str | None) -> str:
    return _OPERATION_STATUS_LABELS.get(operation_status, operation_status or "")


def build_items_xlsx(items: list[dict], sheet_title: str = "库存明细") -> bytes:
    """将单品明细列表生成为 Excel（xlsx）字节流。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = ["商品SN号", "商品名称", "采购单价", "库存状态", "库存属性", "关联单号", "单位组", "单位名称", "关联客户", "操作状态", "备注"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="409EFF")
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for it in items:
        is_in_stock = it.get("stock_status") == "IN_STOCK"
        order_no = it.get("last_order_no") or ""
        unit_price = it.get("unit_price")
        ws.append([
            it.get("item_sn") or "",
            it.get("sku_name") or "",
            float(unit_price) if unit_price is not None else "",
            stock_status_label(it.get("stock_status")),
            stock_condition_label(it.get("stock_status"), it.get("stock_condition")),
            order_no,
            "" if is_in_stock else (it.get("partner_group_name") or ""),
            "" if is_in_stock else (it.get("partner_name") or ""),
            "" if is_in_stock or not order_no.startswith("JOUT-") else (it.get("customer_name") or ""),
            operation_status_label(it.get("operation_status")),
            it.get("remark") or "",
        ])

    widths = [22, 22, 12, 12, 14, 20, 16, 16, 16, 12, 24]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_ledger_xlsx(rows: list[dict]) -> bytes:
    """将库存日流水生成为 Excel（xlsx）字节流。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "库存日流水"

    inbound_headers = LEDGER_INBOUND_DISPLAY_HEADERS
    outbound_headers = LEDGER_OUTBOUND_DISPLAY_HEADERS
    headers = ["日期", "商品 SKU", "期初在库", *inbound_headers, *outbound_headers, "期末在库", "期末资产金额"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="409EFF")
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in rows:
        amount = row.get("closing_asset_amount")
        ws.append([
            row.get("snapshot_date") or "",
            row.get("sku_name") or "",
            row.get("opening_in_stock_qty") if row.get("opening_in_stock_qty") is not None else "",
            *inbound_display_values(row.get("inbound_by_type")),
            *outbound_display_values(row.get("outbound_by_type")),
            row.get("closing_in_stock_qty", 0),
            float(amount) if amount is not None else "",
        ])

    widths = [14, 20, 12, 12, 12, 12, 12, 12, 12, 12, 12, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_ledger_summary_xlsx(rows: list[dict], date_from: str, date_to: str) -> bytes:
    """将按 SKU 区间汇总生成为 Excel（xlsx）字节流。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "库存快照汇总"

    inbound_headers = LEDGER_INBOUND_DISPLAY_HEADERS
    outbound_headers = LEDGER_OUTBOUND_DISPLAY_HEADERS
    headers = [
        "商品 SKU",
        "期初在库",
        *inbound_headers,
        *outbound_headers,
        "理论期末",
        "期末在库",
        "期末资产金额",
        "校验结果",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="409EFF")
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in rows:
        amount = row.get("closing_asset_amount")
        expected = row.get("expected_closing_qty")
        if row.get("balanced"):
            check_result = "账实一致"
        elif expected is not None:
            diff = row.get("diff_qty") or 0
            check_result = f"差异 {diff:+d} 件"
        else:
            check_result = "期初未知"
        ws.append([
            row.get("sku_name") or "",
            row.get("opening_in_stock_qty") if row.get("opening_in_stock_qty") is not None else "",
            *inbound_display_values(row.get("inbound_by_type")),
            *outbound_display_values(row.get("outbound_by_type")),
            expected if expected is not None else "",
            row.get("closing_in_stock_qty", 0),
            float(amount) if amount is not None else "",
            check_result,
        ])

    widths = [20, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 16, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
